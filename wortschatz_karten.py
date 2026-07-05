# -*- coding: utf-8 -*-
"""
Wortschatz Karten - Windows GUI Lernkartenprogramm

Start:
    python wortschatz_karten.py

Abhaengigkeiten:
    pip install openpyxl deep-translator

Die App fragt beim ersten Start nach einer Excel-Datei, speichert die Daten danach
lokal in SQLite und fragt beim naechsten Start nicht erneut. Aktualisierung ueber
"Daten > Excel aktualisieren".
"""

from __future__ import annotations

import os
import re
import sqlite3
import random
import hashlib
import threading
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("Bitte zuerst installieren: pip install openpyxl") from exc

try:
    from deep_translator import GoogleTranslator
except ImportError:
    GoogleTranslator = None

APP_NAME = "Wortschatz Karten"
APP_DIR = Path(os.getenv("APPDATA", str(Path.home()))) / "WortschatzKarten"
DB_PATH = APP_DIR / "wortschatz.db"

STANDARD_COLUMNS = [
    "Unite",
    "Abschnitt",
    "Deutsche Word",
    "Erklärung",
    "Synonym",
    "Antonym",
    "English",
    "Turkish",
    "Beispiel",
]

MUTTERSPRACHEN = {
    "Türkisch": "tr",
    "Englisch": "en",
    "Arabisch": "ar",
    "Russisch": "ru",
    "Persisch": "fa",
    "Spanisch": "es",
    "Französisch": "fr",
    "Italienisch": "it",
    "Polnisch": "pl",
    "Ukrainisch": "uk",
    "Deutsch": "de",
}

# Robust mapping for the Excel variants in the uploaded DSH files.
COLUMN_ALIASES = {
    "unit": ["unite", "unit", "üNITE", "üNİTE", "Ünite", "Unite", "Unit", "teil", "Teil"],
    "abschnitt": ["abschnitt", "bölüm", "bolum", "bereich", "kapitel", "thema", "Teil", "Bölüm"],
    "german_word": ["deutsche word", "almanca kelime", "kelime", "wort", "wort / ausdruck", "kelime / ausdruck", "kelime / ifade", "word / ausdruck", "deutsch", "Deutsch"],
    "erklaerung": ["erklärung", "erklaerung", "almanca açıklama", "almanca aciklama", "deutsche erklärung", "definition"],
    "synonym": ["synonym", "synonym(e)", "synonyme"],
    "antonym": ["antonym", "antonym(e)", "antonyme"],
    "english": ["english", "englisch", "en", "english / tuerkce", "english / türkçe"],
    "turkish": ["turkish", "türkçe", "tuerkce", "tr"],
    "beispiel": ["beispiel", "beispielsatz", "almanca örnek cümle", "almanca ornek cumle", "deutscher beispielsatz"],
}

THEMES = {
    "light": {
        "bg": "#F5F5F7",
        "panel": "#FFFFFF",
        "text": "#1D1D1F",
        "muted": "#6E6E73",
        "line": "#D2D2D7",
        "accent": "#007AFF",
        "accent2": "#34C759",
        "danger": "#FF3B30",
        "chip": "#E9F2FF",
        "button": "#FFFFFF",
        "button_text": "#1D1D1F",
    },
    "dark": {
        "bg": "#161617",
        "panel": "#1F1F22",
        "text": "#F5F5F7",
        "muted": "#A1A1A6",
        "line": "#3A3A3C",
        "accent": "#0A84FF",
        "accent2": "#30D158",
        "danger": "#FF453A",
        "chip": "#26364A",
        "button": "#2C2C2E",
        "button_text": "#F5F5F7",
    },
}


def normalize_header(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.strip().lower()
    text = text.replace("ı", "i").replace("İ", "i")
    text = text.replace("ä", "a").replace("ö", "o").replace("ü", "u").replace("ß", "ss")
    text = re.sub(r"\s+", " ", text)
    text = text.replace("/", " / ")
    text = re.sub(r"\s+", " ", text)
    return text


def text_or_empty(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


@dataclass
class WordCard:
    id: int
    unit: str
    abschnitt: str
    german_word: str
    erklaerung: str
    synonym: str
    antonym: str
    english: str
    turkish: str
    beispiel: str
    known_count: int = 0
    unknown_count: int = 0


class Store:
    def __init__(self, path: Path = DB_PATH) -> None:
        APP_DIR.mkdir(parents=True, exist_ok=True)
        self.path = path
        self.conn = sqlite3.connect(str(path))
        self.conn.row_factory = sqlite3.Row
        self.init_db()

    def init_db(self) -> None:
        cur = self.conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS words (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                unit TEXT NOT NULL,
                abschnitt TEXT NOT NULL,
                german_word TEXT NOT NULL,
                erklaerung TEXT,
                synonym TEXT,
                antonym TEXT,
                english TEXT,
                turkish TEXT,
                beispiel TEXT,
                source_sheet TEXT,
                source_row INTEGER
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS progress (
                word_id INTEGER PRIMARY KEY,
                known_count INTEGER DEFAULT 0,
                unknown_count INTEGER DEFAULT 0,
                last_result TEXT,
                last_seen TEXT,
                FOREIGN KEY(word_id) REFERENCES words(id) ON DELETE CASCADE
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS translations (
                source_text TEXT NOT NULL,
                source_lang TEXT NOT NULL,
                target_lang TEXT NOT NULL,
                translation TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                PRIMARY KEY (source_text, source_lang, target_lang)
            )
        """)
        self.conn.commit()

    def get_setting(self, key: str, default: str = "") -> str:
        row = self.conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
        return row["value"] if row else default

    def set_setting(self, key: str, value: str) -> None:
        self.conn.execute("INSERT OR REPLACE INTO settings(key, value) VALUES(?, ?)", (key, value))
        self.conn.commit()

    def has_words(self) -> bool:
        row = self.conn.execute("SELECT COUNT(*) AS c FROM words").fetchone()
        return bool(row and row["c"] > 0)

    def import_excel(self, excel_path: str) -> int:
        rows = read_excel_words(excel_path)
        if not rows:
            raise ValueError("Keine passenden Wortschatzzeilen gefunden.")
        cur = self.conn.cursor()
        cur.execute("DELETE FROM progress")
        cur.execute("DELETE FROM words")
        for r in rows:
            cur.execute(
                """
                INSERT INTO words(unit, abschnitt, german_word, erklaerung, synonym, antonym, english, turkish, beispiel, source_sheet, source_row)
                VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    r["unit"], r["abschnitt"], r["german_word"], r["erklaerung"],
                    r["synonym"], r["antonym"], r["english"], r["turkish"], r["beispiel"],
                    r.get("source_sheet", ""), r.get("source_row", 0),
                ),
            )
        self.set_setting("source_path", excel_path)
        self.set_setting("last_import", datetime.now().isoformat(timespec="seconds"))
        self.conn.commit()
        return len(rows)

    def units(self) -> List[str]:
        rows = self.conn.execute("SELECT DISTINCT unit FROM words").fetchall()
        values = [r["unit"] for r in rows]
        return sorted(values, key=natural_sort)

    def abschnitte(self, unit: Optional[str] = None) -> List[str]:
        if unit and unit != "Alle":
            rows = self.conn.execute("SELECT DISTINCT abschnitt FROM words WHERE unit = ? ORDER BY abschnitt", (unit,)).fetchall()
        else:
            rows = self.conn.execute("SELECT DISTINCT abschnitt FROM words ORDER BY abschnitt").fetchall()
        return [r["abschnitt"] for r in rows if r["abschnitt"]]

    def query_cards(self, unit: str = "Alle", abschnitt: str = "Alle") -> List[WordCard]:
        sql = """
            SELECT w.*, COALESCE(p.known_count,0) AS known_count, COALESCE(p.unknown_count,0) AS unknown_count
            FROM words w
            LEFT JOIN progress p ON p.word_id = w.id
        """
        params: List[str] = []
        conditions = []
        if unit != "Alle":
            conditions.append("w.unit = ?")
            params.append(unit)
        if abschnitt != "Alle":
            conditions.append("w.abschnitt = ?")
            params.append(abschnitt)
        if conditions:
            sql += " WHERE " + " AND ".join(conditions)
        sql += " ORDER BY w.id"
        rows = self.conn.execute(sql, params).fetchall()
        return [WordCard(**{k: row[k] for k in WordCard.__dataclass_fields__.keys()}) for row in rows]

    def mark(self, word_id: int, result: str) -> None:
        field = "known_count" if result == "known" else "unknown_count"
        now = datetime.now().isoformat(timespec="seconds")
        self.conn.execute(
            f"""
            INSERT INTO progress(word_id, {field}, last_result, last_seen)
            VALUES(?, 1, ?, ?)
            ON CONFLICT(word_id) DO UPDATE SET
                {field} = {field} + 1,
                last_result = excluded.last_result,
                last_seen = excluded.last_seen
            """,
            (word_id, result, now),
        )
        self.conn.commit()

    def get_cached_translation(self, source_text: str, source_lang: str, target_lang: str) -> str:
        row = self.conn.execute(
            """
            SELECT translation FROM translations
            WHERE source_text = ? AND source_lang = ? AND target_lang = ?
            """,
            (source_text, source_lang, target_lang),
        ).fetchone()
        return row["translation"] if row else ""

    def save_cached_translation(self, source_text: str, source_lang: str, target_lang: str, translation: str) -> None:
        if not source_text or not translation:
            return
        self.conn.execute(
            """
            INSERT OR REPLACE INTO translations(source_text, source_lang, target_lang, translation, updated_at)
            VALUES(?, ?, ?, ?, ?)
            """,
            (source_text, source_lang, target_lang, translation, datetime.now().isoformat(timespec="seconds")),
        )
        self.conn.commit()

    def stats(self) -> Dict[str, object]:
        overall = self.conn.execute(
            """
            SELECT COUNT(w.id) AS total,
                   SUM(COALESCE(p.known_count,0)) AS known,
                   SUM(COALESCE(p.unknown_count,0)) AS unknown,
                   SUM(CASE WHEN COALESCE(p.known_count,0)+COALESCE(p.unknown_count,0)>0 THEN 1 ELSE 0 END) AS practiced
            FROM words w LEFT JOIN progress p ON p.word_id = w.id
            """
        ).fetchone()
        by_unit = self.conn.execute(
            """
            SELECT w.unit AS unit, COUNT(w.id) AS total,
                   SUM(COALESCE(p.known_count,0)) AS known,
                   SUM(COALESCE(p.unknown_count,0)) AS unknown,
                   SUM(CASE WHEN COALESCE(p.known_count,0)+COALESCE(p.unknown_count,0)>0 THEN 1 ELSE 0 END) AS practiced
            FROM words w LEFT JOIN progress p ON p.word_id = w.id
            GROUP BY w.unit
            ORDER BY w.unit
            """
        ).fetchall()
        return {"overall": dict(overall), "by_unit": [dict(r) for r in by_unit]}


def natural_sort(text: str) -> Tuple[int, str]:
    nums = re.findall(r"\d+", text or "")
    return (int(nums[0]) if nums else 9999, text or "")


def read_excel_words(excel_path: str) -> List[Dict[str, object]]:
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    all_rows: List[Dict[str, object]] = []

    sheet_titles_norm = {normalize_header(ws.title): ws.title for ws in wb.worksheets}
    preferred_all_sheet = None
    for norm, title in sheet_titles_norm.items():
        if norm in {"tum kelimeler", "tüm kelimeler", "alle worter", "alle wörter", "all words"}:
            preferred_all_sheet = title
            break

    for ws in wb.worksheets:
        title_norm = normalize_header(ws.title)
        if title_norm in {"ozet", "özet", "summary", "zusammenfassung"}:
            continue
        if preferred_all_sheet and ws.title != preferred_all_sheet:
            continue

        header_row_idx, mapping = find_header(ws)
        if not header_row_idx or not mapping:
            continue

        unit_from_sheet = infer_unit_from_sheet(ws.title)
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
            if not row or not any(cell is not None for cell in row):
                continue
            row_values = {}
            for key, col in mapping.items():
                value = row[col - 1] if col - 1 < len(row) else None
                row_values[key] = text_or_empty(value)
            german_word = row_values.get("german_word", "")
            if not german_word or german_word.lower() in {"kelime", "wort", "deutsche word", "almanca kelime"}:
                continue
            unit = row_values.get("unit", "") or unit_from_sheet
            abschnitt = row_values.get("abschnitt", "") or "Ohne Abschnitt"

            # Some combined EN/TR columns may exist in older source files.
            english = row_values.get("english", "")
            turkish = row_values.get("turkish", "")
            if english and not turkish and "/" in english:
                parts = [p.strip() for p in english.split("/", 1)]
                english, turkish = parts[0], parts[1]

            all_rows.append({
                "unit": unit or "Unbekannte Einheit",
                "abschnitt": abschnitt or "Ohne Abschnitt",
                "german_word": german_word,
                "erklaerung": row_values.get("erklaerung", ""),
                "synonym": row_values.get("synonym", ""),
                "antonym": row_values.get("antonym", ""),
                "english": english,
                "turkish": turkish,
                "beispiel": row_values.get("beispiel", ""),
                "source_sheet": ws.title,
                "source_row": row_idx,
            })
    return all_rows


def find_header(ws) -> Tuple[Optional[int], Dict[str, int]]:
    best_row = None
    best_mapping: Dict[str, int] = {}
    max_row = ws.max_row or 15
    max_col = ws.max_column or 20
    for r in range(1, min(max_row, 15) + 1):
        values = [normalize_header(ws.cell(r, c).value) for c in range(1, min(max_col, 20) + 1)]
        mapping: Dict[str, int] = {}
        for field, aliases in COLUMN_ALIASES.items():
            normalized_aliases = {normalize_header(a) for a in aliases}
            for idx, val in enumerate(values, start=1):
                if val in normalized_aliases:
                    mapping[field] = idx
                    break
        # Minimal useful table: word + at least explanation/translation/section.
        score = len(mapping)
        if "german_word" in mapping and score > len(best_mapping):
            best_row, best_mapping = r, mapping
    return best_row, best_mapping


def infer_unit_from_sheet(sheet_name: str) -> str:
    match = re.search(r"(\d+)", sheet_name)
    if match:
        return f"Unit {match.group(1)}"
    return sheet_name.strip() or "Unbekannte Einheit"


def card_icon_text(card: Optional[WordCard]) -> str:
    if not card:
        return "★"
    text = " ".join([card.unit, card.abschnitt, card.german_word]).lower()
    if any(x in text for x in ["geld", "konto", "währung", "waehrung"]):
        return "€"
    if any(x in text for x in ["medizin", "schmerz", "therapie", "chirurgie"]):
        return "+"
    if any(x in text for x in ["kriminal", "gericht", "polizei", "straftat"]):
        return "§"
    if any(x in text for x in ["tier", "haustier", "wildtier", "zoologie"]):
        return "🐾"
    if any(x in text for x in ["chemie", "pestizid", "molekül", "element"]):
        return "⚗"
    if any(x in text for x in ["nachhalt", "ressource", "planet", "umwelt"]):
        return "♻"
    if any(x in text for x in ["sprache", "dialekt", "linguistik", "wort"]):
        return "Aa"
    return (card.german_word[:2] or "★").upper()


def color_for_text(text: str) -> str:
    palette = ["#007AFF", "#5856D6", "#AF52DE", "#FF2D55", "#FF9500", "#34C759", "#5AC8FA"]
    digest = hashlib.md5(text.encode("utf-8")).hexdigest()
    return palette[int(digest[:2], 16) % len(palette)]


class RoundedFrame(tk.Canvas):
    def __init__(self, master, bg_color: str, radius: int = 28, **kwargs):
        super().__init__(master, highlightthickness=0, bd=0, **kwargs)
        self.bg_color = bg_color
        self.radius = radius
        self.bind("<Configure>", self._draw)

    def _draw(self, event=None):
        self.delete("round")
        w, h, r = self.winfo_width(), self.winfo_height(), self.radius
        self.create_round_rect(2, 2, w - 2, h - 2, r, fill=self.bg_color, outline="", tags="round")

    def create_round_rect(self, x1, y1, x2, y2, r, **kwargs):
        points = [
            x1+r, y1, x2-r, y1, x2, y1, x2, y1+r,
            x2, y2-r, x2, y2, x2-r, y2, x1+r, y2,
            x1, y2, x1, y2-r, x1, y1+r, x1, y1,
        ]
        return self.create_polygon(points, smooth=True, **kwargs)


class WortschatzApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.store = Store()
        self.theme_name = self.store.get_setting("theme", "light") or "light"
        self.colors = THEMES[self.theme_name]
        self.cards: List[WordCard] = []
        self.current: Optional[WordCard] = None
        self.is_back = False
        self.deck_index = 0
        self.translation_job_id = 0
        saved_mutter = self.store.get_setting("muttersprache", "Türkisch") or "Türkisch"
        if saved_mutter not in MUTTERSPRACHEN:
            saved_mutter = "Türkisch"
        self.muttersprache_var = tk.StringVar(value=saved_mutter)
        self.title(APP_NAME)
        self.geometry("1080x740")
        self.minsize(960, 650)
        self.configure(bg=self.colors["bg"])
        self.create_menu()
        self.create_widgets()
        self.apply_theme()
        self.after(150, self.ensure_data_loaded)

    def create_menu(self) -> None:
        menubar = tk.Menu(self)
        data_menu = tk.Menu(menubar, tearoff=0)
        data_menu.add_command(label="Excel aktualisieren...", command=self.update_excel_dialog)
        data_menu.add_separator()
        data_menu.add_command(label="Beenden", command=self.destroy)
        menubar.add_cascade(label="Daten", menu=data_menu)

        view_menu = tk.Menu(menubar, tearoff=0)
        view_menu.add_command(label="Light Theme", command=lambda: self.set_theme("light"))
        view_menu.add_command(label="Dark Theme", command=lambda: self.set_theme("dark"))
        menubar.add_cascade(label="Ansicht", menu=view_menu)

        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="Über", command=self.show_about)
        menubar.add_cascade(label="Hilfe", menu=help_menu)
        self.config(menu=menubar)

    def create_widgets(self) -> None:
        self.style = ttk.Style(self)
        self.style.theme_use("clam")

        self.main = tk.Frame(self, bg=self.colors["bg"])
        self.main.pack(fill="both", expand=True, padx=30, pady=24)

        header = tk.Frame(self.main, bg=self.colors["bg"])
        header.pack(fill="x")
        self.title_label = tk.Label(header, text="Wortschatz Karten", font=("Segoe UI Variable", 26, "bold"), bg=self.colors["bg"], fg=self.colors["text"])
        self.title_label.pack(side="left")
        self.subtitle_label = tk.Label(header, text="Minimalistische Lernkarten für DSH-Wortschatz", font=("Segoe UI", 10), bg=self.colors["bg"], fg=self.colors["muted"])
        self.subtitle_label.pack(side="left", padx=(16, 0), pady=(10, 0))

        self.theme_button = tk.Button(header, text="Dark Theme" if self.theme_name == "light" else "Light Theme", command=self.toggle_theme, bd=0, padx=18, pady=9)
        self.theme_button.pack(side="right")
        self.stats_button = tk.Button(header, text="Statistik", command=self.show_stats, bd=0, padx=18, pady=9)
        self.stats_button.pack(side="right", padx=(0, 10))

        controls = tk.Frame(self.main, bg=self.colors["bg"])
        controls.pack(fill="x", pady=(24, 18))
        tk.Label(controls, text="Einheit", bg=self.colors["bg"], fg=self.colors["muted"], font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky="w")
        tk.Label(controls, text="Abschnitt", bg=self.colors["bg"], fg=self.colors["muted"], font=("Segoe UI", 10, "bold")).grid(row=0, column=1, sticky="w", padx=(16, 0))
        tk.Label(controls, text="Muttersprache", bg=self.colors["bg"], fg=self.colors["muted"], font=("Segoe UI", 10, "bold")).grid(row=0, column=2, sticky="w", padx=(16, 0))
        self.unit_var = tk.StringVar(value="Alle")
        self.abschnitt_var = tk.StringVar(value="Alle")
        self.unit_combo = ttk.Combobox(controls, textvariable=self.unit_var, state="readonly", width=26)
        self.abschnitt_combo = ttk.Combobox(controls, textvariable=self.abschnitt_var, state="readonly", width=30)
        self.muttersprache_combo = ttk.Combobox(controls, textvariable=self.muttersprache_var, state="readonly", values=list(MUTTERSPRACHEN.keys()), width=18)
        self.unit_combo.grid(row=1, column=0, sticky="w", pady=(6, 0))
        self.abschnitt_combo.grid(row=1, column=1, sticky="w", padx=(16, 0), pady=(6, 0))
        self.muttersprache_combo.grid(row=1, column=2, sticky="w", padx=(16, 0), pady=(6, 0))
        self.filter_button = tk.Button(controls, text="Filter anwenden", command=self.load_deck, bd=0, padx=18, pady=8)
        self.filter_button.grid(row=1, column=3, sticky="w", padx=(16, 0), pady=(6, 0))
        self.shuffle_button = tk.Button(controls, text="Mischen", command=self.shuffle_deck, bd=0, padx=18, pady=8)
        self.shuffle_button.grid(row=1, column=4, sticky="w", padx=(10, 0), pady=(6, 0))
        self.unit_combo.bind("<<ComboboxSelected>>", lambda e: self.refresh_abschnitte())
        self.muttersprache_combo.bind("<<ComboboxSelected>>", self.on_muttersprache_changed)

        content = tk.Frame(self.main, bg=self.colors["bg"])
        content.pack(fill="both", expand=True)

        self.card_frame = tk.Frame(content, bg=self.colors["panel"], highlightthickness=1, highlightbackground=self.colors["line"])
        self.card_frame.pack(fill="both", expand=True, side="left", padx=(0, 20))

        top_card = tk.Frame(self.card_frame, bg=self.colors["panel"])
        top_card.pack(fill="x", padx=34, pady=(30, 8))
        self.card_meta = tk.Label(top_card, text="", font=("Segoe UI", 11), bg=self.colors["panel"], fg=self.colors["muted"])
        self.card_meta.pack(side="left")
        self.progress_label = tk.Label(top_card, text="", font=("Segoe UI", 11), bg=self.colors["panel"], fg=self.colors["muted"])
        self.progress_label.pack(side="right")

        mid = tk.Frame(self.card_frame, bg=self.colors["panel"])
        mid.pack(fill="both", expand=True, padx=34, pady=10)
        self.illustration = tk.Canvas(mid, width=160, height=160, highlightthickness=0, bg=self.colors["panel"])
        self.illustration.pack(pady=(6, 12))
        self.word_label = tk.Label(mid, text="", font=("Segoe UI Variable", 34, "bold"), wraplength=720, justify="center", bg=self.colors["panel"], fg=self.colors["text"])
        self.word_label.pack(pady=(0, 14))
        self.details_text = tk.Text(mid, height=12, wrap="word", bd=0, font=("Segoe UI", 12), padx=18, pady=18)
        self.details_text.pack(fill="both", expand=True)
        self.details_text.configure(state="disabled")

        actions = tk.Frame(self.card_frame, bg=self.colors["panel"])
        actions.pack(fill="x", padx=34, pady=(12, 30))
        self.flip_button = tk.Button(actions, text="Umdrehen", command=self.flip_card, bd=0, padx=24, pady=12)
        self.flip_button.pack(side="left")
        self.next_button = tk.Button(actions, text="Nächste Karte", command=self.next_card, bd=0, padx=24, pady=12)
        self.next_button.pack(side="left", padx=(10, 0))
        self.known_button = tk.Button(actions, text="Gewusst", command=lambda: self.mark_current("known"), bd=0, padx=24, pady=12)
        self.known_button.pack(side="right")
        self.unknown_button = tk.Button(actions, text="Nicht gewusst", command=lambda: self.mark_current("unknown"), bd=0, padx=24, pady=12)
        self.unknown_button.pack(side="right", padx=(0, 10))

        side = tk.Frame(content, width=250, bg=self.colors["bg"])
        side.pack(fill="y", side="right")
        self.side_panel = tk.Frame(side, bg=self.colors["panel"], highlightthickness=1, highlightbackground=self.colors["line"])
        self.side_panel.pack(fill="both", expand=True)
        tk.Label(self.side_panel, text="Lernmodus", font=("Segoe UI", 17, "bold"), bg=self.colors["panel"], fg=self.colors["text"]).pack(anchor="w", padx=22, pady=(24, 8))
        help_text = (
            "1. Wähle eine Einheit und/oder einen Abschnitt.\n\n"
            "2. Lerne zuerst das deutsche Wort.\n\n"
            "3. Drehe die Karte um und prüfe Erklärung, Synonyme, Übersetzungen, automatische Muttersprache-Übersetzung und Beispiel.\n\n"
            "4. Markiere die Karte als gewusst oder nicht gewusst."
        )
        self.help_label = tk.Label(self.side_panel, text=help_text, justify="left", wraplength=205, font=("Segoe UI", 11), bg=self.colors["panel"], fg=self.colors["muted"])
        self.help_label.pack(anchor="w", padx=22, pady=(4, 20))
        self.status_label = tk.Label(self.side_panel, text="", justify="left", wraplength=205, font=("Segoe UI", 10), bg=self.colors["panel"], fg=self.colors["muted"])
        self.status_label.pack(anchor="w", padx=22, pady=(12, 0))

    def apply_theme(self) -> None:
        c = self.colors
        self.configure(bg=c["bg"])
        for widget in [self.main, self.title_label, self.subtitle_label]:
            try:
                widget.configure(bg=c["bg"])
            except tk.TclError:
                pass
        self.title_label.configure(fg=c["text"])
        self.subtitle_label.configure(fg=c["muted"])
        self.style.configure("TCombobox", fieldbackground=c["panel"], background=c["panel"], foreground=c["text"], arrowcolor=c["text"], bordercolor=c["line"], lightcolor=c["line"], darkcolor=c["line"])
        self.style.map("TCombobox", fieldbackground=[("readonly", c["panel"])] , foreground=[("readonly", c["text"])])

        for frame in self.main.winfo_children():
            if isinstance(frame, tk.Frame):
                try:
                    frame.configure(bg=c["bg"])
                except tk.TclError:
                    pass
        self.card_frame.configure(bg=c["panel"], highlightbackground=c["line"])
        self.side_panel.configure(bg=c["panel"], highlightbackground=c["line"])
        for w in self.card_frame.winfo_children():
            self._theme_recursive(w)
        self._theme_recursive(self.side_panel)
        self.details_text.configure(bg=c["chip"], fg=c["text"], insertbackground=c["text"])
        self.illustration.configure(bg=c["panel"])
        self._style_button(self.theme_button)
        self._style_button(self.stats_button)
        self._style_button(self.filter_button)
        self._style_button(self.shuffle_button)
        self._style_button(self.flip_button, accent=True)
        self._style_button(self.next_button)
        self._style_button(self.known_button, green=True)
        self._style_button(self.unknown_button, danger=True)
        self.theme_button.configure(text="Dark Theme" if self.theme_name == "light" else "Light Theme")
        self.draw_illustration(self.current)

    def _theme_recursive(self, widget) -> None:
        c = self.colors
        try:
            if isinstance(widget, tk.Frame):
                widget.configure(bg=c["panel"])
            elif isinstance(widget, tk.Label):
                widget.configure(bg=c["panel"])
                if widget in {self.card_meta, self.progress_label, self.help_label, self.status_label}:
                    widget.configure(fg=c["muted"])
                else:
                    widget.configure(fg=c["text"])
        except tk.TclError:
            pass
        for child in widget.winfo_children():
            self._theme_recursive(child)

    def _style_button(self, button: tk.Button, accent=False, green=False, danger=False) -> None:
        c = self.colors
        bg = c["button"]
        fg = c["button_text"]
        if accent:
            bg, fg = c["accent"], "#FFFFFF"
        elif green:
            bg, fg = c["accent2"], "#FFFFFF"
        elif danger:
            bg, fg = c["danger"], "#FFFFFF"
        button.configure(bg=bg, fg=fg, activebackground=bg, activeforeground=fg, relief="flat", font=("Segoe UI", 10, "bold"), cursor="hand2")

    def set_theme(self, theme: str) -> None:
        self.theme_name = theme
        self.colors = THEMES[theme]
        self.store.set_setting("theme", theme)
        self.apply_theme()

    def toggle_theme(self) -> None:
        self.set_theme("dark" if self.theme_name == "light" else "light")

    def ensure_data_loaded(self) -> None:
        if not self.store.has_words():
            messagebox.showinfo("Excel-Datei laden", "Bitte wähle beim ersten Start deine Wortschatz-Excel-Datei aus.")
            path = filedialog.askopenfilename(
                title="Wortschatz-Excel auswählen",
                filetypes=[("Excel-Dateien", "*.xlsx *.xlsm"), ("Alle Dateien", "*.*")],
            )
            if not path:
                self.status_label.configure(text="Keine Daten geladen. Bitte über Daten > Excel aktualisieren eine Datei auswählen.")
                return
            try:
                count = self.store.import_excel(path)
                messagebox.showinfo("Import abgeschlossen", f"{count} Wortschatzkarten wurden importiert.")
            except Exception as exc:
                messagebox.showerror("Importfehler", str(exc))
                return
        self.refresh_filters()
        self.load_deck()

    def update_excel_dialog(self) -> None:
        default_path = self.store.get_setting("source_path", "")
        initialdir = str(Path(default_path).parent) if default_path else str(Path.home())
        path = filedialog.askopenfilename(
            title="Wortschatz-Excel aktualisieren",
            initialdir=initialdir,
            filetypes=[("Excel-Dateien", "*.xlsx *.xlsm"), ("Alle Dateien", "*.*")],
        )
        if not path:
            return
        try:
            count = self.store.import_excel(path)
        except Exception as exc:
            messagebox.showerror("Importfehler", str(exc))
            return
        self.refresh_filters()
        self.load_deck()
        messagebox.showinfo("Aktualisiert", f"{count} Karten wurden neu geladen. Der Lernfortschritt wurde zurückgesetzt.")

    def refresh_filters(self) -> None:
        units = ["Alle"] + self.store.units()
        self.unit_combo.configure(values=units)
        if self.unit_var.get() not in units:
            self.unit_var.set("Alle")
        self.refresh_abschnitte()

    def refresh_abschnitte(self) -> None:
        abschnitte = ["Alle"] + self.store.abschnitte(self.unit_var.get())
        self.abschnitt_combo.configure(values=abschnitte)
        if self.abschnitt_var.get() not in abschnitte:
            self.abschnitt_var.set("Alle")

    def load_deck(self) -> None:
        self.cards = self.store.query_cards(self.unit_var.get(), self.abschnitt_var.get())
        random.shuffle(self.cards)
        self.deck_index = -1
        self.status_label.configure(text=f"Aktive Karten: {len(self.cards)}")
        self.next_card()

    def shuffle_deck(self) -> None:
        random.shuffle(self.cards)
        self.deck_index = -1
        self.next_card()

    def next_card(self) -> None:
        if not self.cards:
            self.current = None
            self.word_label.configure(text="Keine Karten gefunden")
            self.card_meta.configure(text="")
            self.progress_label.configure(text="")
            self.set_details("Bitte ändere die Filter oder importiere eine Excel-Datei.")
            self.draw_illustration(None)
            return
        self.deck_index = (self.deck_index + 1) % len(self.cards)
        self.current = self.cards[self.deck_index]
        self.is_back = False
        self.render_card()

    def render_card(self) -> None:
        card = self.current
        if not card:
            return
        self.word_label.configure(text=card.german_word)
        self.card_meta.configure(text=f"{card.unit}  ·  {card.abschnitt}")
        self.progress_label.configure(text=f"{self.deck_index + 1} / {len(self.cards)}")
        self.draw_illustration(card)
        if self.is_back:
            self.set_details(self.format_back(card))
        else:
            self.set_details("Karte umdrehen, um Erklärung, Übersetzungen, Muttersprache und Beispiel zu sehen.")

    def set_details(self, text: str) -> None:
        self.details_text.configure(state="normal")
        self.details_text.delete("1.0", "end")
        self.details_text.insert("1.0", text)
        self.details_text.configure(state="disabled")

    def format_back(self, card: WordCard, word_translation: str = "", example_translation: str = "", loading: bool = False) -> str:
        selected_lang_name = self.muttersprache_var.get()
        if loading:
            word_translation = word_translation or "Übersetzung wird geladen..."
            if card.beispiel:
                example_translation = example_translation or "Übersetzung wird geladen..."
        parts = [
            ("Erklärung", card.erklaerung),
            ("Synonym", card.synonym),
            ("Antonym", card.antonym),
            ("English", card.english),
            ("Türkisch", card.turkish),
            (f"Muttersprache ({selected_lang_name})", word_translation),
            ("Beispiel", card.beispiel),
            (f"Beispiel-Übersetzung ({selected_lang_name})", example_translation),
            ("Fortschritt", f"Gewusst: {card.known_count} · Nicht gewusst: {card.unknown_count}"),
        ]
        return "\n\n".join(f"{label}: {value}" for label, value in parts if value)

    def flip_card(self) -> None:
        self.is_back = not self.is_back
        self.render_card()
        if self.is_back and self.current:
            self.load_muttersprache_translations(self.current)

    def on_muttersprache_changed(self, event=None) -> None:
        self.store.set_setting("muttersprache", self.muttersprache_var.get())
        if self.is_back and self.current:
            self.load_muttersprache_translations(self.current)

    def translate_text_online(self, source_text: str, target_lang: str) -> str:
        source_text = (source_text or "").strip()
        if not source_text:
            return ""
        if target_lang == "de":
            return source_text
        if GoogleTranslator is None:
            return "Übersetzung nicht verfügbar: Bitte installiere deep-translator."
        try:
            return GoogleTranslator(source="de", target=target_lang).translate(source_text) or ""
        except Exception as exc:
            return f"Übersetzung nicht verfügbar: {exc}"

    def load_muttersprache_translations(self, card: WordCard) -> None:
        self.translation_job_id += 1
        job_id = self.translation_job_id
        lang_name = self.muttersprache_var.get()
        lang_code = MUTTERSPRACHEN.get(lang_name, "tr")
        self.store.set_setting("muttersprache", lang_name)

        word_text = card.german_word.strip()
        example_text = card.beispiel.strip()
        cached_word = self.store.get_cached_translation(word_text, "de", lang_code) if word_text else ""
        cached_example = self.store.get_cached_translation(example_text, "de", lang_code) if example_text else ""

        if cached_word and (cached_example or not example_text):
            self.set_details(self.format_back(card, cached_word, cached_example, loading=False))
            return

        self.set_details(self.format_back(card, cached_word, cached_example, loading=True))

        def worker() -> None:
            word_translation = cached_word or self.translate_text_online(word_text, lang_code)
            example_translation = cached_example or self.translate_text_online(example_text, lang_code)
            self.after(0, lambda: self.finish_translation(job_id, card.id, word_text, example_text, lang_code, word_translation, example_translation))

        threading.Thread(target=worker, daemon=True).start()

    def finish_translation(self, job_id: int, card_id: int, word_text: str, example_text: str, lang_code: str, word_translation: str, example_translation: str) -> None:
        if job_id != self.translation_job_id or not self.current or self.current.id != card_id or not self.is_back:
            return
        if word_translation and not word_translation.startswith("Übersetzung nicht verfügbar"):
            self.store.save_cached_translation(word_text, "de", lang_code, word_translation)
        if example_text and example_translation and not example_translation.startswith("Übersetzung nicht verfügbar"):
            self.store.save_cached_translation(example_text, "de", lang_code, example_translation)
        self.set_details(self.format_back(self.current, word_translation, example_translation, loading=False))

    def mark_current(self, result: str) -> None:
        if not self.current:
            return
        self.store.mark(self.current.id, result)
        # Update in-memory counts for immediate feedback.
        if result == "known":
            self.current.known_count += 1
        else:
            self.current.unknown_count += 1
        self.next_card()

    def draw_illustration(self, card: Optional[WordCard]) -> None:
        c = self.colors
        self.illustration.delete("all")
        if not card:
            fill = c["accent"]
            label = "★"
        else:
            fill = color_for_text(card.unit + card.abschnitt + card.german_word)
            label = card_icon_text(card)
        self.illustration.create_oval(18, 18, 142, 142, fill=fill, outline="")
        self.illustration.create_oval(32, 32, 128, 128, fill=self._lighten(fill), outline="")
        self.illustration.create_text(80, 80, text=label, fill="#FFFFFF", font=("Segoe UI Emoji", 34, "bold"))

    def _lighten(self, hex_color: str) -> str:
        hex_color = hex_color.lstrip("#")
        r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
        r = min(255, int(r + (255 - r) * 0.28))
        g = min(255, int(g + (255 - g) * 0.28))
        b = min(255, int(b + (255 - b) * 0.28))
        return f"#{r:02X}{g:02X}{b:02X}"

    def show_stats(self) -> None:
        data = self.store.stats()
        win = tk.Toplevel(self)
        win.title("Statistik")
        win.geometry("760x520")
        win.configure(bg=self.colors["bg"])
        tk.Label(win, text="Lernstatistik", font=("Segoe UI Variable", 24, "bold"), bg=self.colors["bg"], fg=self.colors["text"]).pack(anchor="w", padx=24, pady=(22, 8))
        overall = data["overall"]
        known = overall.get("known") or 0
        unknown = overall.get("unknown") or 0
        total_attempts = known + unknown
        rate = int(round((known / total_attempts) * 100)) if total_attempts else 0
        summary = f"Karten gesamt: {overall.get('total', 0)}   ·   Geübt: {overall.get('practiced', 0)}   ·   Gewusst: {known}   ·   Nicht gewusst: {unknown}   ·   Quote: {rate}%"
        tk.Label(win, text=summary, font=("Segoe UI", 11), bg=self.colors["bg"], fg=self.colors["muted"]).pack(anchor="w", padx=24, pady=(0, 16))

        cols = ("Einheit", "Karten", "Geübt", "Gewusst", "Nicht gewusst", "Quote")
        tree = ttk.Treeview(win, columns=cols, show="headings", height=15)
        for col in cols:
            tree.heading(col, text=col)
            tree.column(col, width=115, anchor="center")
        tree.column("Einheit", width=180, anchor="w")
        for row in data["by_unit"]:
            k = row.get("known") or 0
            u = row.get("unknown") or 0
            attempts = k + u
            q = f"{int(round(k / attempts * 100))}%" if attempts else "-"
            tree.insert("", "end", values=(row["unit"], row["total"], row["practiced"], k, u, q))
        tree.pack(fill="both", expand=True, padx=24, pady=(0, 24))

    def show_about(self) -> None:
        last_import = self.store.get_setting("last_import", "Noch nicht importiert")
        messagebox.showinfo(
            "Über Wortschatz Karten",
            f"{APP_NAME}\n\n"
            "Lernkarten-App für DSH-Wortschatz.\n"
            "Die Daten werden lokal gespeichert.\n\n"
            f"Letzter Import: {last_import}",
        )


def main() -> None:
    app = WortschatzApp()
    app.mainloop()


if __name__ == "__main__":
    main()
